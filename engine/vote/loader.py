"""Vote data XLSX loader (HANDOFF §5.4, §6 control 2)."""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import numpy.typing as npt


@dataclass(frozen=True, slots=True)
class VoteData:
    rankings: npt.NDArray[np.float64]
    entry_ids: tuple[str, ...]
    n_respondents: int


_IMPORTANCE_COL_RE = re.compile(r"^(.+)\s+—\s+Importance(?:\s+\d+)?$")

_ENTRY_NAME_TO_ID: dict[str, str] = {
    "Prompt Injection": "LLM01",
    "Sensitive Information Disclosure": "LLM02",
    "Supply Chain": "LLM03",
    "Supply Chain Vulnerabilities": "LLM03",
    "Data and Model Poisoning": "LLM04",
    "Improper Output Handling": "LLM05",
    "Excessive Agency": "LLM06",
    "System Prompt Leakage": "LLM07",
    "Hidden Context Exposure": "LLM07",
    "Vector and Embedding Weaknesses": "LLM08",
    "Misinformation": "LLM09",
    "Unbounded Consumption": "LLM10",
    "Persistent Memory Poisoning": "NEW-PMP",
    "MCP Tool Interface Exploitation": "NEW-MTIE",
    "Model Misalignment": "NEW-MA",
    "Model Misalignment — 2026 Proposal": "NEW-MA",
    "Inference-Time Side-Channel Disclosure": "NEW-ITSCD",
    "Weaponized LLM Abuse": "NEW-WLA",
    "Model Scheming and Deceptive Alignment": "NEW-MSDA",
    "Cross-Modal Safety Bypass": "ROLL-CMSB",
    "LLM Artifact Promotion Trust Failure": "ROLL-LAPTF",
    "LLM artifact promotion trust failure": "ROLL-LAPTF",
    "Systemic Insecure Code Generation": "ROLL-SICG",
    "Compositional Fine-Tuning Alignment Subversion": "ROLL-CFAS",
    "Compositional Fine-tuning Alignment Subversion": "ROLL-CFAS",
}


def _resolve_entry_id(entry_name: str) -> str | None:
    m = re.match(r"^(LLM\d{2})\b", entry_name)
    if m:
        return m.group(1)
    return _ENTRY_NAME_TO_ID.get(entry_name)


def _importance_to_ranks(scores: npt.NDArray[np.float64]) -> npt.NDArray[np.float64]:
    """Convert importance scores (higher=better) to ranks (1=best) with average tie-breaking."""
    n = len(scores)
    order = np.argsort(-scores)
    ranks = np.empty(n, dtype=np.float64)
    i = 0
    while i < n:
        j = i + 1
        while j < n and scores[order[j]] == scores[order[i]]:
            j += 1
        avg_rank = (i + 1 + j) / 2.0
        for k in range(i, j):
            ranks[order[k]] = avg_rank
        i = j
    return ranks


def load_vote_data(
    xlsx_path: Path,
    sheet_name: str = "Raw Results (Anonymized)",
    column_id_mapping: dict[str, str] | None = None,
) -> VoteData:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {xlsx_path}. "
            f"Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError(f"Sheet '{sheet_name}' is empty")

    header = rows[0]
    has_importance_cols = any(
        h and "— Importance" in str(h) for h in header
    )

    if has_importance_cols:
        return _load_survey_format(header, rows[1:])
    return _load_simple_format(header, rows[1:], column_id_mapping)


def _load_simple_format(
    header: tuple,
    data_rows: list[tuple],
    column_id_mapping: dict[str, str] | None,
) -> VoteData:
    raw_ids = tuple(str(h) for h in header[1:] if h is not None)
    if column_id_mapping:
        entry_ids = tuple(column_id_mapping.get(h, h) for h in raw_ids)
    else:
        entry_ids = raw_ids
    n_entries = len(entry_ids)

    parsed: list[list[float]] = []
    for row in data_rows:
        vals = row[1 : n_entries + 1]
        if all(v is not None for v in vals):
            parsed.append([float(v) for v in vals])

    if not parsed:
        raise ValueError("No respondent data found")

    rankings = np.array(parsed, dtype=np.float64)
    return VoteData(
        rankings=rankings,
        entry_ids=entry_ids,
        n_respondents=rankings.shape[0],
    )


def _load_survey_format(
    header: tuple,
    data_rows: list[tuple],
) -> VoteData:
    importance_cols: list[tuple[int, str]] = []
    for i, h in enumerate(header):
        if h is None:
            continue
        m = _IMPORTANCE_COL_RE.match(str(h))
        if not m:
            continue
        entry_name = m.group(1).strip()
        entry_id = _resolve_entry_id(entry_name)
        if entry_id is not None:
            importance_cols.append((i, entry_id))

    if not importance_cols:
        raise ValueError("No Importance columns found matching taxonomy entries")

    all_entry_ids = sorted(set(eid for _, eid in importance_cols))
    eid_to_idx = {eid: idx for idx, eid in enumerate(all_entry_ids)}
    n_entries = len(all_entry_ids)

    respondent_scores: list[npt.NDArray[np.float64]] = []
    for row in data_rows:
        scores = np.full(n_entries, np.nan)
        for col_idx, entry_id in importance_cols:
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    val = float(row[col_idx])
                    eidx = eid_to_idx[entry_id]
                    if np.isnan(scores[eidx]):
                        scores[eidx] = val
                except (ValueError, TypeError):
                    continue
        if not np.any(np.isnan(scores)):
            respondent_scores.append(scores)

    if not respondent_scores:
        raise ValueError(
            f"No respondents rated all {n_entries} entries. "
            f"Entry IDs: {all_entry_ids}"
        )

    importance_matrix = np.array(respondent_scores, dtype=np.float64)
    n_resp = importance_matrix.shape[0]
    rankings = np.zeros_like(importance_matrix)
    for i in range(n_resp):
        rankings[i] = _importance_to_ranks(importance_matrix[i])

    return VoteData(
        rankings=rankings,
        entry_ids=tuple(all_entry_ids),
        n_respondents=n_resp,
    )
